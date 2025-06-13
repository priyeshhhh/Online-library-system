from django.views.generic import CreateView, ListView, View
from django.urls import reverse_lazy
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
from django.http import HttpResponse
import openpyxl

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'library/form.html'
    success_url = reverse_lazy('author-list')

class AuthorListView(ListView):
    model = Author
    template_name = 'library/list.html'
    context_object_name = 'items'
    paginate_by = 5

class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'library/form.html'
    success_url = reverse_lazy('book-list')

class BookListView(ListView):
    model = Book
    template_name = 'library/list.html'
    context_object_name = 'items'
    paginate_by = 5

class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'library/form.html'
    success_url = reverse_lazy('borrowrecord-list')

class BorrowRecordListView(ListView):
    model = BorrowRecord
    template_name = 'library/list.html'
    context_object_name = 'items'
    paginate_by = 5

class ExportToExcelView(View):
    def get(self, request):
        wb = openpyxl.Workbook()
        del wb['Sheet']

        ws1 = wb.create_sheet(title="Authors")
        ws1.append(['ID', 'Name', 'Email', 'Bio'])
        for a in Author.objects.all():
            ws1.append([a.id, a.name, a.email, a.bio])

        ws2 = wb.create_sheet(title="Books")
        ws2.append(['ID', 'Title', 'Genre', 'Published', 'Author'])
        for b in Book.objects.all():
            ws2.append([b.id, b.title, b.genre, b.published_date, b.author.name])

        ws3 = wb.create_sheet(title="BorrowRecords")
        ws3.append(['ID', 'User', 'Book', 'Borrow Date', 'Return Date'])
        for r in BorrowRecord.objects.all():
            ws3.append([r.id, r.user_name, r.book.title, r.borrow_date, r.return_date])

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
        wb.save(response)
        return response
